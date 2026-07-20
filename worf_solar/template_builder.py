"""
สร้างไฟล์เทมเพลตข้อมูลลูกค้า (.xlsx) ให้ทีมวิเคราะห์/เซลกรอกได้ถูกต้อง
โครงสร้างคอลัมน์ชุดเดียว ใช้ได้ทั้ง 2 กรณี (ตรงกับ data_parser.parse_customer_bill):

  กรณี A — มีข้อมูลรายช่วง (โหลดโปรไฟล์ทุก 15 นาที / รายชั่วโมง)
  กรณี B — มีแค่บิลค่าไฟรายเดือน (PEA/MEA ย้อนหลัง >= 6 เดือน)

เรียก build_template_bytes() ได้จากปุ่มดาวน์โหลดในแอป หรือรันไฟล์นี้ตรงๆ
เพื่อสร้างไฟล์ตัวอย่าง
"""
from __future__ import annotations
import io
import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "ประเภทข้อมูลแถวนี้",          # รายช่วง | รายเดือน
    "วันที่ / เดือน-ปี",            # รายช่วง: 17/01/2568 · รายเดือน: ม.ค. 2568
    "เวลาเริ่มช่วง",                # รายช่วง: 09:15 · รายเดือน: -
    "โหลดเฉลี่ยช่วงนี้ (kW)",       # รายช่วงเท่านั้น
    "หน่วยไฟที่ใช้ (kWh)",          # ใส่ได้ทั้ง 2 กรณี
    "ค่าไฟช่วงนี้ (บาท)",           # รายเดือน: ยอดตามบิล
    "ประเภทช่วง (Peak/Off-Peak)",   # ถ้ามิเตอร์ TOU และบิลแยกให้
    "Demand สูงสุดในช่วง (kW)",     # ถ้าบิลมีค่า Demand
    "ประเภทมิเตอร์",                # TOU / ปกติ / TOD
    "ประเภทธุรกิจ",                 # โรงงาน / สำนักงาน / บ้าน ฯลฯ
    "หมายเหตุ",
]

_ink = Side(style="thin", color="111111")
_border = Border(left=_ink, right=_ink, top=_ink, bottom=_ink)
_head_fill = PatternFill("solid", fgColor="E4D4FB")
_ex_fill = PatternFill("solid", fgColor="FBF1CE")
_head_font = Font(bold=True, size=11)

_TH_MONTH_ABBR = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
                  "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]

# ปัจจัยฤดูกาลอย่างง่าย (ไทย): ร้อนจัด มี.ค.-พ.ค. โหลดสูงสุด (แอร์ทำงานหนัก)
# หนาว พ.ย.-ก.พ. โหลดต่ำสุด ; ฝน มิ.ย.-ต.ค. ปานกลาง — ใช้แค่ให้ตัวอย่างสมจริง ไม่ใช่ค่าอ้างอิง
_SEASONAL_FACTOR = {1: 0.92, 2: 0.95, 3: 1.06, 4: 1.16, 5: 1.12, 6: 1.05,
                    7: 1.02, 8: 1.00, 9: 0.98, 10: 0.97, 11: 0.93, 12: 0.90}


def _style_header(ws, row: int):
    for col in range(1, len(HEADERS) + 1):
        c = ws.cell(row=row, column=col)
        c.font = _head_font
        c.fill = _head_fill
        c.border = _border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, widths=None):
    widths = widths or {}
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 16)


def _day_load_curve(hour_float: float, month: int) -> float:
    """โหลด (kW) ตามชั่วโมงของวัน — โค้งเช้า/เย็นพีค + ปัจจัยฤดูกาล
    รูปแบบเดียวกับโปรไฟล์โรงงาน 2 กะที่ใช้สาธิตในเทมเพลต ไม่ใช่ข้อมูลอ้างอิง
    """
    h = hour_float
    base = 22.0
    morning = 14.0 * math.exp(-((h - 9.0) ** 2) / 8.0)
    evening = 19.0 * math.exp(-((h - 19.0) ** 2) / 6.0)
    dip = -4.0 * math.exp(-((h - 13.0) ** 2) / 3.0)   # พักเที่ยง
    night_floor = 6.0 if h < 5 or h >= 23 else 0.0
    kw = (base + morning + evening + dip) * _SEASONAL_FACTOR.get(month, 1.0) + night_floor
    # จิตเตอร์เล็กน้อยแบบ deterministic (ไม่ใช้ random เพื่อให้ไฟล์สร้างซ้ำได้ค่าเดิมทุกครั้ง)
    jitter = 1.5 * math.sin(h * 2.7 + month * 0.9)
    return max(3.0, round(kw + jitter, 2))


def _interval_example_rows():
    """ตัวอย่างรายช่วง 15 นาที ครบ 12 เดือน — สุ่ม 1 วันตัวแทน/เดือน (วันที่ 15)
    ข้อมูลเต็มวัน (96 ช่วง x 15 นาที) เพื่อให้กราฟ Day/Night ต่อเดือนคำนวณได้ครบ
    (ไฟล์จริงของลูกค้าควรมีข้อมูลทุกวัน ไม่ใช่แค่วันเดียวต่อเดือนแบบตัวอย่างนี้)
    """
    rows = []
    first = True
    for month in range(1, 13):
        date_str = f"15/{month:02d}/2568"
        for q in range(96):  # 00:00 ถึง 23:45 ทีละ 15 นาที
            h_int, m_int = divmod(q * 15, 60)
            time_str = f"{h_int:02d}:{m_int:02d}"
            kw = _day_load_curve(h_int + m_int / 60.0, month)
            on_peak = "Peak" if (9 <= h_int < 22) else "Off-Peak"
            meter_type = "TOU" if first else "-"
            biz_type = "โรงงานผลิต (8-17น.)" if first else "-"
            note = ("ตัวอย่างครบ 12 เดือน (1 วันตัวแทน/เดือน) "
                    "— ไฟล์จริงควรมีข้อมูลทุกวันตลอดช่วงที่มี") if first else ""
            rows.append(["รายช่วง", date_str, time_str, kw, round(kw * 0.25, 2), "-",
                         on_peak, "-", meter_type, biz_type, note])
            first = False
    return rows


def _monthly_example_rows():
    """ตัวอย่างบิลรายเดือน ครบ 12 เดือน (ม.ค.-ธ.ค. 2568) พร้อมความแกว่งตามฤดูกาล"""
    base_kwh, base_rate = 18500.0, 3.70   # บาท/หน่วยเฉลี่ยโดยประมาณ (รวม Ft/demand เฉลี่ยแล้ว)
    rows = []
    for i, month in enumerate(range(1, 13)):
        factor = _SEASONAL_FACTOR.get(month, 1.0)
        kwh = round(base_kwh * factor / 100) * 100          # ปัดให้ดูเป็นตัวเลขบิลจริง
        cost = round(kwh * base_rate / 10) * 10
        month_label = f"{_TH_MONTH_ABBR[month - 1]} 2568"
        rows.append(["รายเดือน", month_label, "-", "-", kwh, cost, "-", "-",
                     "ปกติ (ไม่แยก TOU)" if i == 0 else "-",
                     "โรงงานผลิต (8-17น.)" if i == 0 else "-",
                     "ตัวอย่างครบ 12 เดือน · บิลไม่มีค่า Demand" if i == 0 else ""])
    return rows


def _lead_profile_kw(hour_f: float, month: int, kind: str) -> float:
    """โปรไฟล์โหลดจำลองตามประเภทกิจการ ใช้สร้างไฟล์ตัวอย่างสำหรับสแกนลูกค้า
    kind: 'factory_day' (โรงงานกลางวัน) | 'office' (สำนักงาน) | 'night' (กะกลางคืน)
    """
    h = hour_f
    f = _SEASONAL_FACTOR.get(month, 1.0)
    if kind == "factory_day":
        # เดินเครื่อง 8-17 น. ทุกวัน โหลดกลางวันสูงและนิ่งมาก -> โซลาร์คุ้มสุด (เกรด A)
        base = 45.0
        day = 120.0 if 8 <= h < 17 else 0.0
        lunch = -12.0 * math.exp(-((h - 12.0) ** 2) / 1.2)
        kw = base + day + lunch
    elif kind == "retail":
        # ร้านค้า/ห้าง: เปิด 10:00-21:00 ทุกวัน -> ใช้ไฟช่วงเย็น-ค่ำเยอะพอควร
        # ซึ่งโซลาร์ช่วยไม่ได้ (ไม่มีแดด) -> คุ้มปานกลาง (เกรด B)
        base = 8.0
        openh = 30.0 if 10 <= h < 21 else 0.0
        evening = 25.0 * math.exp(-((h - 19.0) ** 2) / 6.0)   # พีคช่วงค่ำ ไม่มีแดด
        midday = 12.0 * math.exp(-((h - 14.0) ** 2) / 10.0)
        swing = 6.0 * math.sin(h * 2.1 + month)
        kw = base + openh + evening + midday + swing
    else:  # night
        # กะกลางคืน ใช้ไฟตอนไม่มีแดด -> โซลาร์ช่วยได้น้อย (เกรด C)
        base = 20.0
        night = 90.0 if (h >= 19 or h < 6) else 0.0
        kw = base + night
    kw *= f
    kw += 2.0 * math.sin(h * 2.3 + month * 0.7)   # จิตเตอร์เล็กน้อยแบบทำซ้ำได้
    return max(3.0, round(kw, 2))


def _lead_example_rows(kind: str) -> list:
    """ข้อมูลราย 15 นาที ครบ 12 เดือน (1 วันตัวแทน/เดือน x 96 ช่วง) ของลูกค้า 1 ราย"""
    rows = []
    first = True
    for month in range(1, 13):
        date_str = f"15/{month:02d}/2568"
        for q in range(96):
            h_int, m_int = divmod(q * 15, 60)
            kw = _lead_profile_kw(h_int + m_int / 60.0, month, kind)
            biz = {"factory_day": "โรงงานผลิต (8-17น.)", "retail": "ร้านค้า/ห้างสรรพสินค้า",
                   "night": "โรงงานกะกลางคืน"}[kind]
            rows.append(["รายช่วง", date_str, f"{h_int:02d}:{m_int:02d}", kw,
                         round(kw * 0.25, 2), "-",
                         "Peak" if 9 <= h_int < 22 else "Off-Peak", "-",
                         "TOU" if first else "-", biz if first else "-",
                         "ไฟล์ตัวอย่างสำหรับทดสอบสแกนลูกค้า" if first else ""])
            first = False
    return rows


def build_lead_customer_bytes(kind: str) -> bytes:
    """สร้างไฟล์ลูกค้า 1 ราย (.xlsx) สำหรับใช้กับโหมดสแกนลูกค้าหลายราย"""
    wb = Workbook()
    ws = wb.active
    ws.title = "โหลดโปรไฟล์"
    for j, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=j, value=h)
    _style_header(ws, 1)
    for i, r in enumerate(_lead_example_rows(kind), start=2):
        for j, v in enumerate(r, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = _border
    _autosize(ws, {1: 16, 2: 15, 3: 12, 4: 20, 5: 17, 6: 16, 7: 22, 8: 20,
                   9: 15, 10: 20, 11: 22})
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_lead_batch_examples() -> bytes:
    """รวมไฟล์ตัวอย่างลูกค้า 3 ราย (คาดว่าได้เกรดต่างกัน) เป็น .zip ไฟล์เดียว
    เอาไปลากใส่ช่อง 'สแกนหลายรายพร้อมกัน' ได้ทันทีเพื่อทดสอบ
    """
    import zipfile
    files = {
        "ลูกค้า_โรงงานกลางวัน (คาดเกรด A).xlsx": "factory_day",
        "ลูกค้า_ร้านค้า-เปิดถึงค่ำ (คาดเกรด B).xlsx": "retail",
        "ลูกค้า_โรงงานกะกลางคืน (คาดเกรด C).xlsx": "night",
    }
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for fname, kind in files.items():
            z.writestr(fname, build_lead_customer_bytes(kind))
        z.writestr("อ่านก่อน.txt",
                   "ไฟล์ตัวอย่างสำหรับทดสอบโหมด 'หาลูกค้าเชิงรุก — สแกนหลายรายพร้อมกัน'\r\n"
                   "\r\n"
                   "วิธีใช้: แตกไฟล์ zip นี้ แล้วลากไฟล์ .xlsx ทั้ง 3 ไฟล์ใส่ช่องอัปโหลด\r\n"
                   "ในหัวข้อ 'หาลูกค้าเชิงรุก' พร้อมกันได้เลย\r\n"
                   "\r\n"
                   "แต่ละไฟล์ = ลูกค้า 1 ราย (ชื่อไฟล์จะถูกใช้เป็นชื่อลูกค้าในตารางผล)\r\n"
                   "ข้อมูลเป็นราย 15 นาที ครบ 12 เดือน (1 วันตัวแทน/เดือน)\r\n"
                   "\r\n"
                   "ผลที่ควรได้:\r\n"
                   "  โรงงานกลางวัน   -> เกรดสูงสุด (ใช้ไฟกลางวันเยอะ นิ่ง ปริมาณมาก)\r\n"
                   "  ร้านค้าเปิดถึงค่ำ -> เกรดกลาง (ใช้ไฟช่วงค่ำเยอะ โซลาร์ช่วยไม่ได้)\r\n"
                   "  โรงงานกะกลางคืน -> เกรดต่ำสุด (ใช้ไฟตอนไม่มีแดด โซลาร์ช่วยน้อย)\r\n"
                   "\r\n"
                   "ถ้าจะใช้กับลูกค้าจริง: ตั้งชื่อไฟล์เป็นชื่อลูกค้า แล้วกรอกข้อมูล\r\n"
                   "ตามคอลัมน์เดียวกันนี้ (ดูเทมเพลตข้อมูลลูกค้าประกอบ)\r\n")
    return buf.getvalue()


def build_template_bytes() -> bytes:
    wb = Workbook()

    # ---------------- ชีต 1: วิธีใช้ ----------------
    ws = wb.active
    ws.title = "วิธีใช้ (อ่านก่อน)"
    ws.column_dimensions["B"].width = 110
    ws.column_dimensions["A"].width = 4
    lines = [
        ("เทมเพลตข้อมูลลูกค้า — ใช้คอลัมน์ชุดเดียวกันสำหรับทุกกรณี", True),
        ("", False),
        ("ระบบรองรับ 2 กรณี เลือกกรอกตามข้อมูลที่ลูกค้ามีจริง:", True),
        ("  กรณี A) มีโหลดโปรไฟล์รายช่วง (15 นาที/รายชั่วโมง) → กรอกแบบชีต 'กรณี A รายช่วง 15 นาที'", False),
        ("     ใส่ 'รายช่วง' ในคอลัมน์แรกทุกแถว · ต้องมี วันที่ + เวลาเริ่มช่วง + โหลดเฉลี่ยช่วงนี้ (kW)", False),
        ("     → ระบบวิเคราะห์ได้เต็ม: กราฟ Day/Night รายเดือน + เส้นแนวโน้ม, Demand Peak,", False),
        ("        ผลประหยัด 'จำลองจากโหลดจริง' รายแพ็กเกจ", False),
        ("  กรณี B) มีแค่บิลค่าไฟรายเดือน (PEA/MEA) ย้อนหลังอย่างน้อย 6 เดือน → กรอกแบบชีต 'กรณี B บิลรายเดือน'", False),
        ("     ใส่ 'รายเดือน' ในคอลัมน์แรกทุกแถว · ต้องมี เดือน-ปี + หน่วยไฟ (kWh) + ค่าไฟ (บาท)", False),
        ("     → ระบบวิเคราะห์ได้บางส่วน: แนวโน้มหน่วยไฟ/ค่าไฟ + ผลประหยัดแบบ 'ช่วงประมาณการ'", False),
        ("        (สิ่งที่ทำไม่ได้จะถูกขีดฆ่าแสดงไว้บนหน้าเว็บ พร้อมเหตุผล)", False),
        ("", False),
        ("กติกาสำคัญ (ถ้าไม่ทำ ระบบอาจอ่านไฟล์ผิด):", True),
        ("  1. ห้ามเปลี่ยนชื่อหัวคอลัมน์ · แถวหนึ่ง = ช่วงเวลาหนึ่ง (หรือหนึ่งเดือน)", False),
        ("  2. ช่องที่ไม่มีข้อมูลใส่ '-' หรือเว้นว่าง อย่าใส่ 0 แทนค่าที่ไม่รู้", False),
        ("  3. ปี พ.ศ. หรือ ค.ศ. ก็ได้ ระบบแปลงให้อัตโนมัติ (เช่น 17/01/2568 หรือ 2025-01-17)", False),
        ("  4. ตอนอัปโหลดจริง ให้เหลือชีตข้อมูลเพียงชีตเดียว (ลบชีตตัวอย่างที่ไม่ใช้และชีตวิธีใช้นี้ทิ้ง)", False),
        ("  5. ตัวเลขห้ามมีหน่วยปนในช่อง เช่น ใส่ 18500 ไม่ใช่ '18,500 kWh'  (ใส่ comma ได้ ระบบตัดให้)", False),
        ("  6. กรณี A ถ้ามีข้อมูลหลายเดือนยิ่งดี — กราฟแนวโน้ม Day/Night ต้องการอย่างน้อย 2 เดือน "
         "(ตัวอย่างในไฟล์นี้ใส่ครบ 12 เดือนให้ดูเป็นแนวทาง)", False),
        ("", False),
        ("มุมเซล (คนขาย): ยิ่งขอโหลดโปรไฟล์รายช่วงจากลูกค้าได้ ตัวเลขที่โชว์จะเป็น 'จำลองจากการใช้ไฟจริง'", False),
        ("ซึ่งน่าเชื่อถือกว่าตัวเลขแคตตาล็อก และระบบจะโชว์เทียบกันทั้งสองแบบให้ลูกค้าเห็นภาพ", False),
        ("มุมนักวิเคราะห์: ถ้าได้ไฟล์มิเตอร์ราย 15 นาทีของหน่วยงาน ให้แปลงเป็นรูปแบบชีต 'กรณี A' ก่อนอัปโหลด", False),
    ]
    for i, (txt, bold) in enumerate(lines, start=2):
        c = ws.cell(row=i, column=2, value=txt)
        c.font = Font(bold=bold, size=12 if bold else 11)
        c.alignment = Alignment(wrap_text=False)

    # ---------------- ชีต 2: กรณี A รายช่วง ----------------
    wa = wb.create_sheet("กรณี A รายช่วง 15 นาที")
    wa.cell(row=1, column=1, value="ตัวอย่างกรอก กรณี A — โหลดโปรไฟล์รายช่วง 15 นาที "
            "ครบ 12 เดือน (สุ่ม 1 วันตัวแทน/เดือน x ข้อมูลเต็มวันทุก 15 นาที = 96 ช่วง/วัน) "
            "— ไฟล์จริงของลูกค้าควรมีข้อมูลทุกวันตลอดช่วงที่มี ไม่ใช่แค่วันเดียวต่อเดือนแบบตัวอย่างนี้"
            ).font = Font(bold=True)
    hdr_row = 3
    for j, h in enumerate(HEADERS, start=1):
        wa.cell(row=hdr_row, column=j, value=h)
    _style_header(wa, hdr_row)
    for i, r in enumerate(_interval_example_rows(), start=hdr_row + 1):
        for j, v in enumerate(r, start=1):
            c = wa.cell(row=i, column=j, value=v)
            c.border = _border
            c.fill = _ex_fill
    _autosize(wa, {1: 16, 2: 15, 3: 12, 4: 20, 5: 17, 6: 16, 7: 22, 8: 20, 9: 15, 10: 20, 11: 22})
    wa.freeze_panes = f"A{hdr_row + 1}"

    # ---------------- ชีต 3: กรณี B บิลรายเดือน ----------------
    wbm = wb.create_sheet("กรณี B บิลรายเดือน")
    wbm.cell(row=1, column=1, value="ตัวอย่างกรอก กรณี B — บิลค่าไฟรายเดือน ครบ 12 เดือน "
             "(ม.ค.-ธ.ค. 2568) ย้อนหลังอย่างน้อย 6 เดือนใช้งานได้ แต่ครบปีจะเห็นแนวโน้มตามฤดูกาลชัดกว่า"
             ).font = Font(bold=True)
    hdr_row = 3
    for j, h in enumerate(HEADERS, start=1):
        wbm.cell(row=hdr_row, column=j, value=h)
    _style_header(wbm, hdr_row)
    for i, r in enumerate(_monthly_example_rows(), start=hdr_row + 1):
        for j, v in enumerate(r, start=1):
            c = wbm.cell(row=i, column=j, value=v)
            c.border = _border
            c.fill = _ex_fill
    _autosize(wbm, {1: 16, 2: 15, 3: 12, 4: 20, 5: 17, 6: 16, 7: 22, 8: 20, 9: 18, 10: 20, 11: 22})
    wbm.freeze_panes = f"A{hdr_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "เทมเพลตข้อมูลลูกค้า_ตัวอย่าง.xlsx"
    with open(out, "wb") as f:
        f.write(build_template_bytes())
    print("saved:", out)
